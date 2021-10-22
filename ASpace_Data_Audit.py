import mysql.connector as mysql
import re
from asnake.client import ASnakeClient
from mysql.connector import errorcode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from secrets import *

id_field_regex = re.compile(r"(^id_+\d)")
id_combined_regex = re.compile(r'[\W_]+', re.UNICODE)

client = ASnakeClient(baseurl=as_api, username=as_un, password=as_pw)
client.authorize()


def connect_db():
    try:
        staging_connect = mysql.connect(user=as_dbstag_un,
                                        password=as_dbstag_pw,
                                        host=as_dbstag_host,
                                        database=as_dbstag_database,
                                        port=as_dbstag_port)
    except mysql.Error as error:
        if error.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your username or password")
        elif error.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(error)
    else:
        print(staging_connect)
        staging_cursor = staging_connect.cursor()
        return staging_connect, staging_cursor


def query_database(connection, cursor, statement):
    cursor.execute(statement)
    results = cursor.fetchall()
    cursor.close()
    connection.close()
    return results


def generate_spreadsheet():
    wb = Workbook()
    data_spreadsheet = 'reports/data_audit.xlsx'
    wb.save(data_spreadsheet)
    return wb, data_spreadsheet


def standardize_resids(results):
    redone_results = []
    for result in list(results):
        new_result = list(result)
        res_id_list = new_result[1].strip('[').strip(']').replace('"', '').split(',')
        combined_id = ''
        for id_comp in res_id_list:
            if id_comp != 'null':
                combined_id += id_comp + '-'
        new_result[1] = combined_id[:-1]
        redone_results.append(new_result)
    return redone_results


def update_booleans(results):
    checked_results = []
    for result in list(results):
        single_row = list(result)
        list_index = 0
        for value in single_row:
            if value == 0:
                single_row[list_index] = False
            elif value == 1:
                single_row[list_index] = True
            list_index += 1
        checked_results.append(single_row)
    return checked_results


def run_query(wb, sheetname, headers, statement, resid=False, booleans=False):
    tc_nobar = wb.create_sheet(sheetname)
    tc_nobar.title = sheetname
    header_index = 0
    for row in tc_nobar.iter_rows(min_row=1, max_col=len(headers)):
        for cell in row:
            tc_nobar[cell.coordinate] = headers[header_index]
            tc_nobar[cell.coordinate].font = Font(bold=True, underline='single')
            header_index += 1
    connection, cursor = connect_db()
    results = query_database(connection, cursor, statement)
    standardized_results = results
    if resid is True:
        standardized_results = standardize_resids(results)
    if booleans is True:
        standardized_results = update_booleans(standardized_results)
    for result in standardized_results:
        tc_nobar.append(result)


def check_controlled_vocabs(wb, terms, vocab, vocab_num):
    vocab_sheet = wb.create_sheet(f"{vocab}")
    vocab_sheet.title = f"{vocab}"
    write_row_index = 2
    headers = [f"{vocab}", "Read Only?", "Suppressed?"]
    header_index = 0
    for row in vocab_sheet.iter_rows(min_row=1, max_col=3):
        for cell in row:
            vocab_sheet[cell.coordinate] = headers[header_index]
            vocab_sheet[cell.coordinate].font = Font(bold=True, underline='single')
            header_index += 1
    statement = (f'SELECT CONVERT(ev.value using utf8) AS {vocab}, ev.readonly AS Read_Only, '
                 f'ev.suppressed AS Suppressed '
                 f'FROM enumeration_value AS ev '
                 f'WHERE enumeration_id = {vocab_num}')
    connection, cursor = connect_db()
    standardize_results = update_booleans(query_database(connection, cursor, statement))
    for result in standardize_results:
        vocab_sheet.append(result)
        if result[0] not in terms:
            for cell in vocab_sheet[f'{write_row_index}:{write_row_index}']:
                cell.fill = PatternFill(start_color='FFFF0000',
                                        end_color='FFFF0000',
                                        fill_type='solid')
        write_row_index += 1


def check_child_levels(tree_info, child_counts, root_uri, aspace_coll_id, client, top_level=False):
    levels = []
    children = []
    if tree_info["uri"] not in child_counts:
        child_counts[f"{tree_info['uri']}"] = (tree_info["title"], tree_info["level"], aspace_coll_id)
        print(aspace_coll_id)
    if "precomputed_waypoints" in tree_info and tree_info["child_count"] != 0:
        if top_level is True:
            waypoint_key = ""
        else:
            waypoint_key = tree_info["uri"]
        for waypoint_num, waypoint_info in tree_info["precomputed_waypoints"][waypoint_key].items():
            for child in waypoint_info:
                if child["level"] not in levels:
                    levels.append(child["level"])
                    child_counts[f'{child["uri"]}'] = (child["title"], child["child_count"], child["level"],
                                                       aspace_coll_id)
                print(" " * 10 + f'Checking {child["title"]}')
                children = client.get(root_uri + "/tree/node", params={"node_uri": child["uri"],
                                                                       "published_only": True}).json()
                # check_child_levels(children, child_counts, root_uri, aspace_coll_id, client, top_level=False)
    try:  # TODO: figure out a way to maintain levels for just the same level
        if not children:
            return child_counts
    finally:
        if children:
            check_child_levels(children, child_counts, root_uri, aspace_coll_id, client, top_level=False)


def check_ao_levels():
    child_levels = {}
    repos = client.get("repositories").json()
    for repo in repos:
        print(repo["name"] + "\n")
        repo_id = repo["uri"].split("/")[2]
        resources = client.get("repositories/{}/resources".format(repo_id), params={"all_ids": True}).json()
        for resource_id in resources:
            resource = client.get("repositories/{}/resources/{}".format(repo_id, resource_id))
            combined_id = ""
            for field, value in resource.json().items():
                id_match = id_field_regex.match(field)
                if id_match:
                    combined_id += value + "-"
            combined_id = combined_id[:-1]
            combined_aspace_id_clean = id_combined_regex.sub('', combined_id)
            if resource.json()["publish"] is True:
                if resource.status_code == 200:
                    root_uri = f'/repositories/{repo_id}/resources/{resource_id}'
                    tree_info = client.get(f'{root_uri}/tree/root').json()
                    print(combined_id)
                    child_levels = check_child_levels(tree_info, child_levels, root_uri, combined_id, client,
                                                      top_level=True)
    return child_levels


def run():
    workbook, spreadsheet = generate_spreadsheet()
    controlled_vocabs = {"Subject_Term_Type": [["cultural_context", "function", "genre_form", "geographic",
                                                "occupation", "style_period", "technique", "temporal", "topical",
                                                "uniform_title"], 54],
                         "Subject_Sources": [["aat", "lcsh", "local", "tgn", "lcnaf"], 23],
                         "Finding_Aid_Status_Terms": [["completed", "unprocessed", "in_process", "problem"], 21],
                         "Name_Sources": [["local", "naf", "ingest"], 4],
                         "Instance_Types": [["audio", "books", "digital_object", "graphic_materials", "maps",
                                             "microform", "mixed_materials", "moving_images", "electronic_records",
                                             "artifacts"], 22],
                         "Extent_Types": [["linear_feet", "box(es)", "item(s)", "gigabyte(s)", "pages", "volume(s)",
                                           "moving_image(s)", "interview(s)", "minutes", "folder(s)",
                                           "sound_recording(s)", "photograph(s)", "oversize_folders"], 14],
                         "Digital_Object_Types": [["cartographic", "mixed_materials", "moving_image",
                                                   "software_multimedia", "sound_recording", "still_image", "text"],
                                                  12],
                         "Container_Types": [["box", "folder", "oversized_box", "oversized_folder", "reel", "roll",
                                              "portfolio", "item", "volume", "physdesc", "electronic_records", "carton",
                                              "drawer", "cassette", "rr", "cs"], 16],
                         "Accession_Resource_Types": [["collection", "papers", "records"], 7]}
    for term, info in controlled_vocabs.items():
        check_controlled_vocabs(workbook, info[0], term, info[1])
    cuid_statement = ('SELECT repo.name AS Repository, resource.identifier AS Resource_ID, ao.ref_id AS Ref_ID, '
                      'ao.title AS Archival_Object_Title, ao.component_id AS Component_Unique_Identifier '
                      'FROM archival_object AS ao '
                      'JOIN repository AS repo ON repo.id = ao.repo_id '
                      'JOIN resource ON resource.id = ao.root_record_id '
                      'WHERE component_id is not Null '
                      'AND resource.publish is True '
                      'AND ao.publish is True')
    tcnb_statement = ('SELECT repo.name AS Repository, resource.identifier AS Resource_ID, '
                      'ao.ref_id AS Linked_Archival_Object_REFID, ao.title AS Linked_Archival_Object_Title, '
                      'top_container.indicator, CONVERT(ev.value using utf8) AS Container_Type '
                      'FROM top_container '
                      'JOIN top_container_link_rlshp AS top_rlsh ON top_rlsh.top_container_id = top_container.id '
                      'JOIN sub_container ON top_rlsh.sub_container_id = sub_container.id '
                      'JOIN instance ON instance.id = sub_container.instance_id '
                      'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                      'JOIN repository AS repo ON repo.id = ao.repo_id '
                      'JOIN resource ON resource.id = ao.root_record_id '
                      'JOIN enumeration_value AS ev ON ev.id = top_container.type_id '
                      'WHERE top_container.barcode is NULL AND ao.publish is True AND resource.publish is True')
    tcnind_statement = ('SELECT repo.name AS Repository, resource.identifier AS Resource_ID, '
                        'ao.ref_id AS Linked_Archival_Object_REFID, ao.title AS Linked_Archival_Object_Title, '
                        'top_container.indicator, CONVERT(ev.value using utf8) AS Container_Type '
                        'FROM top_container '
                        'JOIN top_container_link_rlshp AS top_rlsh ON top_rlsh.top_container_id = top_container.id '
                        'JOIN sub_container ON top_rlsh.sub_container_id = sub_container.id '
                        'JOIN instance ON instance.id = sub_container.instance_id '
                        'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                        'JOIN repository AS repo ON repo.id = ao.repo_id '
                        'JOIN resource ON resource.id = ao.root_record_id '
                        'JOIN enumeration_value AS ev ON ev.id = top_container.type_id '
                        'WHERE top_container.indicator is NULL AND ao.publish is True AND resource.publish is True')
    users_statement = ('SELECT name, username, is_system_user AS System_Administrator, is_hidden_user AS Hidden_User '
                       'FROM user')
    aomtc_statement = ('SELECT repository.name AS Repository, instance.archival_object_id, ao.ref_id, ao.title, '
                       'COUNT(*) AS Top_Container_Count '
                       'FROM instance '
                       'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                       'JOIN repository ON repository.id = ao.repo_id '
                       'WHERE ao.publish is True AND instance.instance_type_id != 349 '
                       'group by instance.archival_object_id HAVING count(archival_object_id) > 1')
    aomdo_statement = ('SELECT repository.name AS Repository, instance.archival_object_id, ao.ref_id, ao.title, '
                       'COUNT(*) AS Digital_Object_Count '
                       'FROM instance '
                       'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                       'JOIN repository ON repository.id = ao.repo_id '
                       'WHERE ao.publish is True AND instance.instance_type_id = 349 '
                       'group by instance.archival_object_id HAVING count(archival_object_id) > 1')
    aocollevel_statement = ("SELECT repository.name AS Repository, resource.identifier AS Resource_ID, ao.title, "
                            "ao.ref_id, ev.value "
                            "FROM archival_object AS ao "
                            "JOIN resource ON ao.root_record_id = resource.id "
                            "JOIN repository ON ao.repo_id = repository.id "
                            "JOIN enumeration_value AS ev ON ao.level_id = ev.id "
                            "WHERE ao.parent_id is Null "
                            "AND ao.publish = 1 "
                            "AND resource.publish is True "
                            "AND ao.level_id = 889")
    eadid_statement = ("SELECT repository.name, resource.title, resource.identifier, resource.ead_id "
                       "FROM resource "
                       "JOIN repository ON repository.id = resource.repo_id "
                       "WHERE resource.ead_id is not NULL AND resource.publish is TRUE")
    queries = {"Component Unique Identifiers": [["Repository", "Resource ID", "RefID", "Archival Object Title",
                                                 "Component Unique Identifier"], cuid_statement, {"resids": True},
                                                {"booleans": False}],
               "Top Containers - No Barcodes": [["Repository", "Resource ID", "RefID", "Archival Object Title",
                                                 "Top Container Indicator", "Container Type"], tcnb_statement,
                                                {"resids": True}, {"booleans": False}],
               "Top Containers - No Indicator": [["Repository", "Resource ID", "RefID", "Archival Object Title",
                                                 "Top Container Indicator", "Container Type"], tcnind_statement,
                                                {"resids": True}, {"booleans": False}],
               "Users": [["Name", "Username", "System Administrator?", "Hidden User?"], users_statement,
                         {"resids": False}, {"booleans": True}],
               "Arch Objs-Multiple Top Conts": [["Repository", "Archival Object ID", "RefID",
                                                 "Archival Object title", "Top Container Count"],
                                                aomtc_statement, {"resids": False}, {"booleans": False}],
               "Arch Objs-Multiple Dig Objs": [["Repository", "Archival Object ID", "RefID",
                                                "Archival Object Title", "Digital Object Count"],
                                               aomdo_statement, {"resids": False}, {"booleans": False}],
               "Arch Objs-Collection Level": [["Repository", "Resource ID", "Archival Object Title", "RefID",
                                               "Level"], aocollevel_statement, {"resids": True}, {"booleans": False}],
               "EAD-IDs": [["Repository", "Resource Title", "Resource ID", "EAD ID"], eadid_statement,
                           {"resids": True}, {"booleans": False}]}
    for query, info in queries.items():
        run_query(workbook, query, info[0], info[1], resid=info[2]["resids"], booleans=info[3]["booleans"])

    workbook.remove(workbook["Sheet"])
    workbook.save(spreadsheet)


run()